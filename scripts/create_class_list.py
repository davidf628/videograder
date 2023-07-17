import csv, os, sys
from openpyxl import load_workbook

sys.path.append(os.path.join('..', 'lib'))
import lib.util as util

config = util.load_config('config.toml')

ONE_DRIVE = config(['onedrive'])
CLASS_SCHEDULE_FILE = config(['class_schedule_file'])

OUTPUT_FILE = 'classes_new.csv'

termdates = { '2023U' : ['5/25/2023', '8/10/2023'],
              '2023UF1' : ['5/25/2023', '6/30/2023'],
              '2023UF2' : ['7/5/2023', '8/10/2023'],
              '2023UM' : ['5/1/2023', '5/25/2023'] }

videosets = { 'mat-101' : 'isp',
              'mat-102' : 'isp',
              'mat-033' : 'mat033',
              'mat-110' : 'mat110',
              'mat-111' : 'mat111',
              'mat-112' : 'mat112',
			  'mat-115' : 'mat115',
              'mat-120' : 'mat120',
              'mat-130' : 'mat130',
              'mat-140' : 'mat140',
              'mat-141' : 'mat141' }

def main():
    class_records = load_schedule(os.path.join(ONE_DRIVE, CLASS_SCHEDULE_FILE))
    create_record_file(OUTPUT_FILE, class_records)
                                

def create_record_file(output_file, class_records):
    for i in class_records:
        print(i)

    # Write class_records to a csv file
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(class_records)

def load_schedule(schedule_file):
    
    class_records = [ ['videoset', 'course', 'termstart', 'termend', 'instructor', 'email'] ]
    current_course_number = ''
    course = ''
    section = ''
    semester = ''
    instructor = ''
    email = ''

    # Read the course workbook
    wb = load_workbook(schedule_file)
    ws = wb.active

    # Go through line by line and determine class parameters like:
    for row in range(2, ws.max_row):

        # Each section number for a class should be in column B
        section = ws.cell(row=row, column=2).value

        # section == None for blank lines in workbook, skip those
        if section != None:

            # Frank leaves off padded zeros for some reason
            section = pad_zeros(str(section))

            # course number is in column A
            course_number = ws.cell(row=row, column=1).value

            # Preserve the course number used because it's only indicated
            #   once on the spreadsheet
            if course_number == None:
                course_number = current_course_number
            else:
                if course_number == '101/102':
                    course_number = '101-102'
                current_course_number = course_number

            course = f'mat-{course_number}'

            if course in videosets.keys():

                # term data is in column C
                semester = ws.cell(row=row, column=3).value

                # name of instructor is in column E
                instructor = ws.cell(row=row, column=5).value

                if instructor == None:
                    instructor = 'nobody'
                else:
                    instructor = instructor.replace("'", '')
                    instructor = instructor.replace(' ', '')
                    instructor = instructor.replace('.', '')
            
                # Video set to use
                videoset = videosets[course].lower()

                # Course title
                course_title = f'{course}-{section}'
                
                # Term start date
                termstart = termdates[semester][0]

                # Term end date
                termend = termdates[semester][1]

                # Instructor Name
                instructor_name = f'{instructor}_{course_number}_{section}_{semester}'.lower()

                # Now create a record with all this information
                record = [ videoset, course_title, termstart, termend, instructor_name, email ]
                
                # And add the new record to a list
                class_records.append(record)
                #print(instructor_name)

    return class_records


def pad_zeros(section):
    if section == None:
        return ''
    else:
        while len(section) < 3:
            if section[0] in [ 'W', 'P', 'D', 'B', 'Y' ]:
                ch = section[0]
                section = section.replace(ch, '')
                section = ch + '0' + section
            else:
                section = '0' + section
        return section.lower()


if __name__ == '__main__':
    main()
